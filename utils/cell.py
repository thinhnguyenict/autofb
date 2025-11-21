import openpyxl

def update_excel(excel_file, col, row):
    wb = openpyxl.load_workbook(excel_file)
        # Select the active sheet
    sheet = wb.active
    print(sheet.title)
    sheet[f"{col}{row}"] = "KKK"
    wb.save(excel_file)


def read_excel(excel_file):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file)
        # Select the active sheet
        sheet = wb.active
        
        # Initialize empty lists to store data and positions
        messages = []
        links = []
        paths = []
        publish_status = []
        publish_status_positions = []

        # # Iterate over the first column in the sheet
        # for row_idx, row in enumerate(sheet.iter_cols(min_col=1, max_col=1, values_only=True), start=1):
        #     # Remove extra whitespace from cell value and convert to lowercase
        #     cleaned_cell_value = row[0].strip().lower() if isinstance(row[0], str) else row[0]
        #     # Check if the cell contains the value 'publish_status'
        #     if cleaned_cell_value == 'publish_status':
        #         # Return the row number
        #         return row_idx

        # Iterate over rows and columns to extract data and positions
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            cleaned_row = [col.strip().lower() for col in row]
            # Extract data from each row
            message = row[0]  # Assuming message is in the first column
            link = row[1]     # Assuming link is in the second column
            path = row[2]     # Assuming path is in the third column
            status = row[3]
            publish_status_position = cleaned_row.index('publish_status') + 1
            
            import ipdb; ipdb.set_trace()
            # Append data to corresponding lists
            messages.append(message)
            links.append(link)
            paths.append(path)
            publish_status.append(status)          
            publish_status_positions.append(('D', publish_status_position))
            
        return messages, links, paths, publish_status_positions
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")
        return None, None, None, None


messages, links, paths, message_positions, link_positions, path_positions, publish_status_positions = read_excel("./data.xlsx")

print(messages, links, paths, message_positions, link_positions, path_positions, publish_status_positions )
print(publish_status_positions)
# update_excel("./data.xlsx", "D", "2")
